import random
import os
import sys
import pickle
import pdb

import numpy as np 
import pandas as pd

from collections import Counter, defaultdict
from openpyxl import load_workbook
from pprint import pprint

sys.path.insert(1, './lib/pyNMR')
import nmrDataMod as ndm


def stratified_group_k_fold(x, y, groups, k, seed=None):
    label_count = np.max(y) + 1
    label_counts_per_group = defaultdict(lambda: np.zeros(label_count))
    label_distribution = Counter()
    for label, group in zip(y, groups):
        label_counts_per_group[group][label] += 1
        label_distribution[label] += 1
    
    label_counts_per_fold = defaultdict(lambda: np.zeros(label_count))
    groups_per_fold = defaultdict(set)

    def eval_label_counts_per_fold(label_counts, fold):
        label_counts_per_fold[fold] += label_counts
        std_per_label = []
        for label in range(label_count):
            label_std = np.std([label_counts_per_fold[i][label] / label_distribution[label] for i in range(k)])
            std_per_label.append(label_std)
        label_counts_per_fold[fold] -= label_counts
        return np.mean(std_per_label)
    
    groups_and_label_counts = list(label_counts_per_group.items())
    random.Random(seed).shuffle(groups_and_label_counts)

    for group, label_counts in sorted(groups_and_label_counts, key=lambda x: -np.std(x[1])):
        best_fold = None
        min_eval = None
        for i in range(k):
            fold_eval = eval_label_counts_per_fold(label_counts, i)
            if min_eval is None or fold_eval < min_eval:
                min_eval = fold_eval
                best_fold = i
        label_counts_per_fold[best_fold] += label_counts
        groups_per_fold[best_fold].add(group)

    all_groups = set(groups)
    for i in range(k):
        train_groups = all_groups - groups_per_fold[i]
        test_groups = groups_per_fold[i]

        train_indices = [i for i, group in enumerate(groups) if group in train_groups]
        test_indices = [i for i, group in enumerate(groups) if group in test_groups]

        yield train_indices, test_indices

def get_distribution(labels):
    label_distribution = Counter(labels)
    sum_labels = sum(label_distribution.values())
    return [f'{label_distribution[i] / sum_labels:.2%}' for i in range(np.max(labels) + 1)]

def generate_datasets(dataset):
    # control vs agressive
    count = 0
    with open("./data/folds/control_aggressive.txt", "w") as f:
        idx = 1
        for patient in dataset:
            flag = False
            for data in patient:
                path = './data/neurosurgery_glioma/' + str(data[0]) + '/4/'
                if not (os.path.exists(path)):
                    continue
                with open(path + "pulseprogram", 'r') as f1:
                    experiment = f1.readlines()[1].translate(str.maketrans('', '', ';')).strip()
                if (experiment == "cpmgpr1d"): # only cpmg experiments
                    if data[4] == "Agressive-GLIOMA":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 1, data[3], float(data[2])))
                        count += 1
                    elif data[4] == "Control":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 0, data[3], float(data[2])))
                        count += 1
            if flag == True:
                idx += 1
    print("Control-Agressive Data Count: ", count)
    
    # benign vs agressive
    count = 0
    with open("./data/folds/benign_aggressive.txt", "w") as f:
        idx = 1
        for patient in dataset:
            flag = False
            for data in patient:
                path = './data/neurosurgery_glioma/' + str(data[0]) + '/4/'
                if not (os.path.exists(path)):
                    continue
                with open(path + "pulseprogram", 'r') as f1:
                    experiment = f1.readlines()[1].translate(str.maketrans('', '', ';')).strip()
                if (experiment == "cpmgpr1d"): # only cpmg experiments
                    if data[4] == "Agressive-GLIOMA":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 1, data[3], float(data[2])))
                        count += 1
                    elif data[4] == "Benign-GLIOMA":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 0, data[3],float(data[2])))
                        count += 1
            if flag == True:
                idx += 1
    print("Benign Agressive Data Count: ", count)

    # control vs tumor(benign + agressive)
    count = 0
    with open("./data/folds/control_tumor.txt", "w") as f:
        idx = 1
        for patient in dataset:
            flag = False
            for data in patient:
                path = './data/neurosurgery_glioma/' + str(data[0]) + '/4/'
                if not (os.path.exists(path)):
                    continue
                with open(path + "pulseprogram", 'r') as f1:
                    experiment = f1.readlines()[1].translate(str.maketrans('', '', ';')).strip()
                if (experiment == "cpmgpr1d"): # only cpmg experiments
                    if data[4] == "Agressive-GLIOMA" or data[4] == "Benign-GLIOMA":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 1, data[3], float(data[2])))
                        count += 1
                    elif data[4] == "Control":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 0, data[3], float(data[2])))
                        count += 1
            if flag == True:
                idx += 1
    print("Control Tumor Data Count: ", count)

    # multiclass
    count = 0
    with open("./data/folds/multiclass.txt", "w") as f:
        idx = 1
        for patient in dataset:
            flag = False
            for data in patient:
                path = './data/neurosurgery_glioma/' + str(data[0]) + '/4/'
                if not (os.path.exists(path)):
                    continue
                with open(path + "pulseprogram", 'r') as f1:
                    experiment = f1.readlines()[1].translate(str.maketrans('', '', ';')).strip()
                if (experiment == "cpmgpr1d"): # only cpmg experiments
                    if data[4] == "Agressive-GLIOMA":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 1, data[3], float(data[2])))
                        count += 1
                    elif data[4] == "Control":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 0, data[3], float(data[2])))
                        count += 1
                    elif data[4] == "Benign-GLIOMA":
                        flag = True
                        f.write("%d,%s,%d,%s,%f\n" % (idx, data[0], 2, data[3], float(data[2])))
                        count += 1
            if flag == True:
                idx += 1
    print("Control Benign Aggressive Data Count: ", count)

# process free induction decay(fid) data and create a spectrum similar to Topspin's output
def preprocess_spectrum(path):
    data = ndm.nmrData(path, "TopSpin")
    shiftPoints = 70  # first 70 points are not actual data.
    data.leftShift(0, 1, shiftPoints)
    # data.lineBroadening(1,2,10)
    data.fourierTransform(1,2)
    phase = data.autoPhase0(2,0,-50000,50000)
    data.phase(2,3, phase)
    return np.absolute(data.allFid[3][0])

# add spectrum to the corresponding dataset
def add_spectra_to_dataset(dataset_path, n, slice_start, slice_end, H_filename, labels_filename, patients_filename, pathology_classes_filename, nrmn_filename):

    print("Length of a single data:", slice_end-slice_start)

    # read data
    data = pd.read_csv(dataset_path, header=None)
    patient_ids = data.values[:, 0]
    filenames = data.values[:, 1]
    gt_labels = data.values[:, 2]
    pathology_classes = data.values[:, 3]
    masses = data.values[:, 4]
    
    # generate training and test samples
    H = np.zeros([slice_end - slice_start, n])
    labels = np.zeros([n, 1])
    patients = np.zeros([n, 1])
    p_classes = np.zeros([n, 1], dtype='object')
    nrmn = np.zeros([n, 1], dtype='object')

    # read the file second time
    counter = 0
    for idx, name in enumerate(filenames):
        path = './data/neurosurgery_glioma/' + str(name) + '/4/'
        if not (os.path.exists(path)):
            continue
        with open(path + "pulseprogram", 'r') as f:
            experiment = f.readlines()[1].translate(str.maketrans('', '', ';')).strip()
        if (experiment == "cpmgpr1d"): # only cpmg experiments
            # normalize each sample to have a mass of 20mg
            H[:,counter] = preprocess_spectrum(path)[slice_start:slice_end] #* 20 / masses[idx]
            labels[counter] = gt_labels[idx]
            patients[counter] = patient_ids[idx]
            p_classes[counter] = pathology_classes[idx]
            nrmn[counter] = str(name)
            counter += 1
    

    with open("./data/folds/" + str(patients_filename),'wb') as f:
        pickle.dump(patients,f)
    with open("./data/folds/" + str(H_filename),'wb') as f:
        pickle.dump(H,f)
    with open("./data/folds/" + str(labels_filename),'wb') as f:
        pickle.dump(labels,f)
    with open("./data/folds/" + str(pathology_classes_filename),'wb') as f:
        pickle.dump(p_classes,f)
    with open("./data/folds/" + str(nrmn_filename),'wb') as f:
        pickle.dump(nrmn,f)

#----------------------- dataset augmentation pipeline --------------------------

# first open the xlsx file 
data_path = './data/'
csv_path = 'Neurosurgery_january_2020.xlsx'
path = os.path.join(data_path, csv_path)

# open xlsx file and work in Feuil1 sheet
wb = load_workbook(path)
ws = wb["Feuil1"]

# get N.RMN(a unique id), Groupe, pathology_class, pathology_label, localization
# the results of TEST group patients appear at pathology_label and the corresponding pathology_label is empty
n_rmn, masses, patient_groups, pathology_results, pathology_labels, tumor_locations  = ws["D"], ws["E"], ws["F"], ws["G"], ws["H"], ws["N"]

# --------------------------generate dataset------------------
dataset = []
outer_idx = -1
prev_identifier = ""
for index, (id, group, mass, p_result, p_label, loc) in enumerate(zip(n_rmn, patient_groups, masses, pathology_results, pathology_labels, tumor_locations)):
    # skip first two lines 
    if index <= 1:
        continue
    # dont continue processing after EOF
    if index >= 778:
        break
    # process N.RMN to determine the patients of the data
    trimmed_id = id.value[4:]
    parts = trimmed_id.split('-')
    if len(parts[0]) <= 1:
        identifier = parts[1]
    else:
        identifier = parts[0]
    if identifier != prev_identifier:
        outer_idx += 1
        dataset.append([])
        prev_pathology_label = ""
    if group.value == "GLIOMA":
        prev_pathology_label = p_label.value
        dataset[outer_idx].append((id.value, group.value, mass.value, p_result.value, p_label.value, loc.value))
    elif group.value == "CONTROL":
        dataset[outer_idx].append((id.value, group.value, mass.value, p_result.value, p_label.value, loc.value))
    elif group.value == "TEST":
        if p_result.value == "pos":
            if p_label.value is None:
                dataset[outer_idx].append((id.value, group.value, mass.value, p_result.value, prev_pathology_label, loc.value))
            else: 
                dataset[outer_idx].append((id.value, group.value, mass.value, p_result.value, p_label.value, loc.value))
                prev_pathology_label = p_label.value
        elif p_result.value == "neg":
            dataset[outer_idx].append((id.value, group.value, mass.value, p_result.value, "Control", loc.value))    
    prev_identifier = identifier

generate_datasets(dataset)

# create of the selected dataset
# uncomment corresponding line to generate
# dataset_name = "control_aggressive"  
# dataset_name = "control_tumor"
dataset_name = "benign_aggressive"  
# dataset_name = "multiclass"

# random seed for reproducibility
# the previous seed was 1000
SEED = np.random.randint(low=0, high=3000)
SEED = 2251
print("SEED: ", SEED)

print(dataset_name)
dataset_path = f"./data/folds/{dataset_name}.txt"
data = pd.read_csv(dataset_path, header=None)
patient_ids = data.values[:, 0]
filenames = data.values[:, 1]
labels = data.values[:, 2]
pathology_labels = data.values[:, 3]
masses = data.values[:, 4]

label_distributions = [get_distribution(labels)]
index = ["dataset"]

prev_test_patients = []
fold_size = []
for fold, (train_idx, test_idx) in enumerate(stratified_group_k_fold(filenames.tolist(), labels.tolist(), patient_ids.tolist(), k=8, seed=SEED)):
    train_x, test_x = filenames[train_idx], filenames[test_idx]
    train_y, test_y = labels[train_idx], labels[test_idx]
    train_groups, test_groups = patient_ids[train_idx], patient_ids[test_idx]
    train_pathology_labels, test_pathology_labels = pathology_labels[train_idx], pathology_labels[test_idx]
    train_masses, test_masses = masses[train_idx], masses[test_idx]
    assert len(set(train_groups) & set(test_groups)) == 0
    assert len(set(test_groups) & set(prev_test_patients)) == 0
    label_distributions.append(get_distribution(test_y))
    index.append(f'fold{fold}')
    prev_test_patients = test_groups
    fold_size.append(test_groups.shape[0])
    with open("./data/folds/"+ f"{dataset_name}_fold_{fold}" + ".txt", "w") as f:
        for idx in range(test_groups.shape[0]):
            f.write("%d,%s,%d,%s,%f\n" % (test_groups[idx], test_x[idx], test_y[idx], test_pathology_labels[idx], float(test_masses[idx])))

'''
    PCA PVE > 0.98 features are between the following indices:
    Control-Tumor: 1 - 8172
    Control-Benign: 4 - 8166
    Benign-Agressive: 2 - 8174
    Control-Agressive 1 - 8171
'''
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_0.txt", fold_size[0], 0, 8172, f"{dataset_name}_hydrogens_fold_1", f"{dataset_name}_labels_fold_1", f"{dataset_name}_patients_1", f"{dataset_name}_pathology_classes_1", f"{dataset_name}_nrmn_1")
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_1.txt", fold_size[1], 0, 8172, f"{dataset_name}_hydrogens_fold_2", f"{dataset_name}_labels_fold_2", f"{dataset_name}_patients_2", f"{dataset_name}_pathology_classes_2", f"{dataset_name}_nrmn_2")
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_2.txt", fold_size[2], 0, 8172, f"{dataset_name}_hydrogens_fold_3", f"{dataset_name}_labels_fold_3", f"{dataset_name}_patients_3", f"{dataset_name}_pathology_classes_3", f"{dataset_name}_nrmn_3")
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_3.txt", fold_size[3], 0, 8172, f"{dataset_name}_hydrogens_fold_4", f"{dataset_name}_labels_fold_4", f"{dataset_name}_patients_4", f"{dataset_name}_pathology_classes_4", f"{dataset_name}_nrmn_4")
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_4.txt", fold_size[4], 0, 8172, f"{dataset_name}_hydrogens_fold_5", f"{dataset_name}_labels_fold_5", f"{dataset_name}_patients_5", f"{dataset_name}_pathology_classes_5", f"{dataset_name}_nrmn_5")
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_5.txt", fold_size[5], 0, 8172, f"{dataset_name}_hydrogens_fold_6", f"{dataset_name}_labels_fold_6", f"{dataset_name}_patients_6", f"{dataset_name}_pathology_classes_6", f"{dataset_name}_nrmn_6")
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_6.txt", fold_size[6], 0, 8172, f"{dataset_name}_hydrogens_fold_7", f"{dataset_name}_labels_fold_7", f"{dataset_name}_patients_7", f"{dataset_name}_pathology_classes_7", f"{dataset_name}_nrmn_7")
add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_7.txt", fold_size[7], 0, 8172, f"{dataset_name}_hydrogens_fold_8", f"{dataset_name}_labels_fold_8", f"{dataset_name}_patients_8", f"{dataset_name}_pathology_classes_8", f"{dataset_name}_nrmn_8")
# add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_8.txt", fold_size[8], 2, 8174, f"{dataset_name}_hydrogens_fold_9", f"{dataset_name}_labels_fold_9", f"{dataset_name}_patients_9", f"{dataset_name}_pathology_classes_9", f"{dataset_name}_nrmn_9")
# add_spectra_to_dataset( f"./data/folds/{dataset_name}_fold_9.txt", fold_size[9], 2, 8174, f"{dataset_name}_hydrogens_fold_10", f"{dataset_name}_labels_fold_10", f"{dataset_name}_patients_10", f"{dataset_name}_pathology_classes_10", f"{dataset_name}_nrmn_10")



print('Distributions per class:')
print(pd.DataFrame(label_distributions, index= index, columns=[f'Label {l}' for l in range(np.max(train_y) + 1)]))



